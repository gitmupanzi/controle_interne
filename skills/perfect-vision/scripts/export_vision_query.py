from __future__ import annotations

import argparse
import re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import pyodbc
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from profile_vision_queries import QUERIES, extract_queries


ROOT = Path(__file__).resolve().parents[3]


def safe_filename(value: str) -> str:
    normalized = re.sub(r"[^0-9A-Za-zÀ-ÿ_-]+", "_", value).strip("_")
    return normalized[:120] or "requete"


def excel_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, bytes):
        return value.hex()
    if isinstance(value, str) and len(value) > 32_767:
        return value[:32_767]
    return value


def build_batch(query_sql: str, args: argparse.Namespace) -> str:
    devise = "NULL" if args.id_devise is None else str(args.id_devise)
    return f"""
SET NOCOUNT ON;
SET XACT_ABORT ON;
SET LOCK_TIMEOUT {args.lock_timeout_ms};
DECLARE @date_debut date = '{args.date_start}';
DECLARE @date_fin date = '{args.date_end}';
DECLARE @seuil_5k_usd_cdf float = {args.threshold_5k_cdf};
DECLARE @seuil_10k_usd_cdf float = {args.threshold_10k_cdf};
DECLARE @id_devise_reporting int = {devise};
DECLARE @taux_usd_cdf decimal(19,6) = {args.usd_cdf_rate};
DECLARE @convertir_affichage_cdf bit = 1;
{query_sql}
"""


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Exécute une requête autonome du catalogue Perfect Vision et exporte son résultat en XLSX."
    )
    parser.add_argument("--query", type=int, required=True, help="Numéro de la requête du catalogue.")
    parser.add_argument("--server", default="localhost")
    parser.add_argument("--database", default="BB_VISION_PRO_TEST")
    parser.add_argument("--date-start", default="2026-06-01")
    parser.add_argument("--date-end", default="2026-06-30")
    parser.add_argument("--id-devise", type=int, choices=[1, 2], default=None)
    parser.add_argument("--threshold-5k-cdf", type=float, default=11_375_000)
    parser.add_argument("--threshold-10k-cdf", type=float, default=22_750_000)
    parser.add_argument("--usd-cdf-rate", type=float, default=2_275)
    parser.add_argument("--query-timeout", type=int, default=600)
    parser.add_argument("--lock-timeout-ms", type=int, default=30_000)
    parser.add_argument("--fetch-size", type=int, default=2_000)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()

    queries = {query["number"]: query for query in extract_queries(QUERIES)}
    if args.query not in queries:
        raise SystemExit(f"Requête inconnue : {args.query}")
    query = queries[args.query]

    output = args.output
    if output is None:
        output = (
            ROOT
            / "reports"
            / "sql_exports"
            / (
                f"{args.query:03d}_{safe_filename(query['title'])}_"
                f"{args.date_start}_{args.date_end}.xlsx"
            )
        )
    output = output.resolve()
    output.parent.mkdir(parents=True, exist_ok=True)

    connection_string = (
        "DRIVER={ODBC Driver 17 for SQL Server};"
        f"SERVER={args.server};DATABASE={args.database};"
        "Trusted_Connection=yes;TrustServerCertificate=yes;"
    )

    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("Résultats")
    worksheet.freeze_panes = "A2"
    row_count = 0

    with pyodbc.connect(connection_string, autocommit=True) as connection:
        connection.timeout = args.query_timeout
        cursor = connection.cursor()
        cursor.execute(build_batch(query["sql"], args))
        while cursor.description is None:
            if not cursor.nextset():
                raise RuntimeError(f"La requête {args.query} ne retourne aucun jeu de résultats.")

        columns = [str(column[0]) for column in cursor.description]
        worksheet.append(columns)
        while True:
            rows = cursor.fetchmany(args.fetch_size)
            if not rows:
                break
            for row in rows:
                worksheet.append([excel_value(value) for value in row])
            row_count += len(rows)

    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{row_count + 1}"
    workbook.save(output)
    print(
        f"Q{args.query:03d} exportée : {row_count} ligne(s), "
        f"{len(columns)} colonne(s), fichier={output}"
    )


if __name__ == "__main__":
    main()
